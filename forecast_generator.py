"""予算・FC・実績を横並びで比較するExcelを生成するスクリプト."""

from __future__ import annotations

import datetime
import re
import tkinter as tk
from tkinter import filedialog, simpledialog

import jpholiday
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def find_date_column(df: pd.DataFrame) -> str:
    """Return the column name that represents a date.

    The function normalizes column names by stripping whitespace and converting
    them to lowercase. It then searches for a column likely representing a
    date, such as "日付", "宿泊日" or "date". If no suitable column is found,
    a ``KeyError`` is raised.
    """

    normalized = {c.strip().lower(): c for c in df.columns}
    candidates = ["日付", "宿泊日", "date"]

    for key, original in normalized.items():
        cleaned = re.sub(r"\s+", "", key)
        if cleaned in candidates:
            return original

    for key, original in normalized.items():
        if re.search(r"日\s*付", key) or "宿泊日" in key or "date" in key:
            return original

    raise KeyError("日付に該当する列が見つかりません")

# === GUIでキャパシティと期首月の取得 ===

try:
    root = tk.Tk()
    root.withdraw()
    capacity = simpledialog.askinteger(
        "キャパシティ入力",
        "客室キャパシティ（部屋数）を入力してください：",
    )
    start_month = simpledialog.askinteger(
        "期首月入力",
        "期首月（例：1〜12）を入力してください：",
    )
    file_path = filedialog.askopenfilename(title="日別予算Excelファイルを選択")
except tk.TclError:
    print("GUI を起動できないため CLI モードで実行します。")
    capacity = int(input("キャパシティ（部屋数）: "))
    start_month = int(input("期首月 (1-12): "))
    file_path = input("日別予算Excelファイルのパス: ")
# === Excel読み込み ===
xls = pd.read_excel(file_path, sheet_name=None)

# === 出力用Excel作成 ===
wb = Workbook()
wb.remove(wb.active)

# === 曜日装飾 ===
sat_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
sat_font = Font(color="003366")
sun_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
sun_font = Font(color="990000")
gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
budget_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
fc_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
thin = Side(style="thin", color="999999")
medium = Side(style="medium", color="999999")

# --- 各月シート作成 ---
summary_dict: dict[tuple[int, int], dict[str, int]] = {}
for sheet_name, df in xls.items():
    if not isinstance(df, pd.DataFrame) or df.empty:
        continue

    try:
        date_col = find_date_column(df)
    except KeyError:
        continue
    df["日付"] = pd.to_datetime(df[date_col], errors="coerce")
    df = df.dropna(subset=["日付"]).sort_values("日付")
    if df.empty:
        continue

    year = int(df["日付"].dt.year.iloc[0])
    month = int(df["日付"].dt.month.iloc[0])
    ws = wb.create_sheet(title=f"{year}年{month}月")

    # 日付ごとに横持ち化
    metrics = ["室数", "人数", "宿泊売上", "OCC", "ADR", "DOR", "RevPAR"]
    rows: list[dict] = []
    for _, r in df.iterrows():
        date = r["日付"].date()
        weekday_jp = ["月", "火", "水", "木", "金", "土", "日"][date.weekday()]
        if jpholiday.is_holiday(date):
            weekday_jp = f"{weekday_jp}・祝"

        base = {
            "日付": date.strftime("%Y/%m/%d"),
            "曜日": weekday_jp,
            "室数_予算": r.get("室数", ""),
            "人数_予算": r.get("人数", ""),
            "宿泊売上_予算": r.get("宿泊売上", ""),
        }
        for m in ["OCC", "ADR", "DOR", "RevPAR"]:
            base[f"{m}_予算"] = ""
        for kind in ["FC", "OH", "実績"]:
            for m in metrics:
                base[f"{m}_{kind}"] = ""
        rows.append(base)

    df_out = pd.DataFrame(rows)

    # 列順を定義
    metrics = ["室数", "人数", "宿泊売上", "OCC", "ADR", "DOR", "RevPAR"]
    cols = ["日付", "曜日"]
    cols += [f"{m}_予算" for m in metrics]
    cols += [f"{m}_FC" for m in metrics]
    cols += ["差_OCC_FC-予算", "差_ADR_FC-予算", "差_売上_FC-予算"]
    cols += [f"{m}_OH" for m in metrics]
    cols += ["差_OCC_OH-FC", "差_ADR_OH-FC", "差_売上_OH-FC"]
    cols += [f"{m}_実績" for m in metrics]
    cols += ["差_OCC_実績-FC", "差_ADR_実績-FC", "差_売上_実績-FC"]
    for c in cols:
        if c not in df_out.columns:
            df_out[c] = ""
    df_out = df_out.reindex(columns=cols)

    # 出力
    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    header_map = {c.value: i for i, c in enumerate(ws[1], start=1)}
    ws.freeze_panes = "C2"
    data_end_row = ws.max_row
    for row in range(2, data_end_row + 1):
        # 予算列の指標をExcel数式で計算
        room_c = get_column_letter(header_map["室数_予算"])
        pax_c = get_column_letter(header_map["人数_予算"])
        sales_c = get_column_letter(header_map["宿泊売上_予算"])
        occ_c = ws.cell(row=row, column=header_map["OCC_予算"])
        adr_c = ws.cell(row=row, column=header_map["ADR_予算"])
        dor_c = ws.cell(row=row, column=header_map["DOR_予算"])
        rev_c = ws.cell(row=row, column=header_map["RevPAR_予算"])

        occ_c.value = f"={room_c}{row}/{capacity}"
        adr_c.value = f"={sales_c}{row}/{room_c}{row}"
        dor_c.value = f"={pax_c}{row}/{room_c}{row}"
        rev_c.value = f"={sales_c}{row}/{capacity}"

        occ_c.number_format = "0.0%"
        adr_c.number_format = "#,##0"
        dor_c.number_format = "0.00"
        rev_c.number_format = "#,##0"
        for col_name in ["室数_予算", "人数_予算", "宿泊売上_予算"]:
            ws.cell(row=row, column=header_map[col_name]).number_format = "#,##0"

        # FC列の数式
        fc_room_c = get_column_letter(header_map["室数_FC"])
        fc_pax_c = get_column_letter(header_map["人数_FC"])
        fc_sales_c = get_column_letter(header_map["宿泊売上_FC"])
        fc_occ = ws.cell(row=row, column=header_map["OCC_FC"])
        fc_adr = ws.cell(row=row, column=header_map["ADR_FC"])
        fc_dor = ws.cell(row=row, column=header_map["DOR_FC"])
        fc_rev = ws.cell(row=row, column=header_map["RevPAR_FC"])

        fc_occ.value = (
            f"=IF({fc_room_c}{row}=\"\", \"\", {fc_room_c}{row}/{capacity})"
        )
        fc_adr.value = (
            f"=IF(OR({fc_sales_c}{row}=\"\", {fc_room_c}{row}=\"\"), \"\", {fc_sales_c}{row}/{fc_room_c}{row})"
        )
        fc_dor.value = (
            f"=IF(OR({fc_pax_c}{row}=\"\", {fc_room_c}{row}=\"\"), \"\", {fc_pax_c}{row}/{fc_room_c}{row})"
        )
        fc_rev.value = (
            f"=IF({fc_sales_c}{row}=\"\", \"\", {fc_sales_c}{row}/{capacity})"
        )

        fc_occ.number_format = "0.0%"
        fc_adr.number_format = "#,##0"
        fc_dor.number_format = "0.00"
        fc_rev.number_format = "#,##0"

        # OH列の数式
        oh_room_c = get_column_letter(header_map["室数_OH"])
        oh_pax_c = get_column_letter(header_map["人数_OH"])
        oh_sales_c = get_column_letter(header_map["宿泊売上_OH"])
        oh_occ = ws.cell(row=row, column=header_map["OCC_OH"])
        oh_adr = ws.cell(row=row, column=header_map["ADR_OH"])
        oh_dor = ws.cell(row=row, column=header_map["DOR_OH"])
        oh_rev = ws.cell(row=row, column=header_map["RevPAR_OH"])

        oh_occ.value = (
            f"=IF({oh_room_c}{row}=\"\", \"\", {oh_room_c}{row}/{capacity})"
        )
        oh_adr.value = (
            f"=IF(OR({oh_sales_c}{row}=\"\", {oh_room_c}{row}=\"\"), \"\", {oh_sales_c}{row}/{oh_room_c}{row})"
        )
        oh_dor.value = (
            f"=IF(OR({oh_pax_c}{row}=\"\", {oh_room_c}{row}=\"\"), \"\", {oh_pax_c}{row}/{oh_room_c}{row})"
        )
        oh_rev.value = (
            f"=IF({oh_sales_c}{row}=\"\", \"\", {oh_sales_c}{row}/{capacity})"
        )

        oh_occ.number_format = "0.0%"
        oh_adr.number_format = "#,##0"
        oh_dor.number_format = "0.00"
        oh_rev.number_format = "#,##0"

        # 実績列の数式
        act_room_c = get_column_letter(header_map["室数_実績"])
        act_pax_c = get_column_letter(header_map["人数_実績"])
        act_sales_c = get_column_letter(header_map["宿泊売上_実績"])
        act_occ = ws.cell(row=row, column=header_map["OCC_実績"])
        act_adr = ws.cell(row=row, column=header_map["ADR_実績"])
        act_dor = ws.cell(row=row, column=header_map["DOR_実績"])
        act_rev = ws.cell(row=row, column=header_map["RevPAR_実績"])

        act_occ.value = (
            f"=IF({act_room_c}{row}=\"\", \"\", {act_room_c}{row}/{capacity})"
        )
        act_adr.value = (
            f"=IF(OR({act_sales_c}{row}=\"\", {act_room_c}{row}=\"\"), \"\", {act_sales_c}{row}/{act_room_c}{row})"
        )
        act_dor.value = (
            f"=IF(OR({act_pax_c}{row}=\"\", {act_room_c}{row}=\"\"), \"\", {act_pax_c}{row}/{act_room_c}{row})"
        )
        act_rev.value = (
            f"=IF({act_sales_c}{row}=\"\", \"\", {act_sales_c}{row}/{capacity})"
        )

        act_occ.number_format = "0.0%"
        act_adr.number_format = "#,##0"
        act_dor.number_format = "0.00"
        act_rev.number_format = "#,##0"

        # 手入力セルの表示形式
        for col_name in [
            "室数_FC",
            "人数_FC",
            "宿泊売上_FC",
            "室数_OH",
            "人数_OH",
            "宿泊売上_OH",
            "室数_実績",
            "人数_実績",
            "宿泊売上_実績",
        ]:
            ws.cell(row=row, column=header_map[col_name]).number_format = "#,##0"

        # 差異列
        ws.cell(row=row, column=header_map["差_OCC_FC-予算"]).value = (
            f"=IF({get_column_letter(header_map['OCC_FC'])}{row}=\"\", \"\", {get_column_letter(header_map['OCC_FC'])}{row}-{get_column_letter(header_map['OCC_予算'])}{row})"
        )
        ws.cell(row=row, column=header_map["差_ADR_FC-予算"]).value = (
            f"=IF({get_column_letter(header_map['ADR_FC'])}{row}=\"\", \"\", {get_column_letter(header_map['ADR_FC'])}{row}-{get_column_letter(header_map['ADR_予算'])}{row})"
        )
        ws.cell(row=row, column=header_map["差_売上_FC-予算"]).value = (
            f"=IF({get_column_letter(header_map['宿泊売上_FC'])}{row}=\"\", \"\", {get_column_letter(header_map['宿泊売上_FC'])}{row}-{get_column_letter(header_map['宿泊売上_予算'])}{row})"
        )
        ws.cell(row=row, column=header_map["差_OCC_OH-FC"]).value = (
            f"=IF({get_column_letter(header_map['OCC_OH'])}{row}=\"\", \"\", {get_column_letter(header_map['OCC_OH'])}{row}-{get_column_letter(header_map['OCC_FC'])}{row})"
        )
        ws.cell(row=row, column=header_map["差_ADR_OH-FC"]).value = (
            f"=IF({get_column_letter(header_map['ADR_OH'])}{row}=\"\", \"\", {get_column_letter(header_map['ADR_OH'])}{row}-{get_column_letter(header_map['ADR_FC'])}{row})"
        )
        ws.cell(row=row, column=header_map["差_売上_OH-FC"]).value = (
            f"=IF({get_column_letter(header_map['宿泊売上_OH'])}{row}=\"\", \"\", {get_column_letter(header_map['宿泊売上_OH'])}{row}-{get_column_letter(header_map['宿泊売上_FC'])}{row})"
        )
        ws.cell(row=row, column=header_map["差_OCC_実績-FC"]).value = (
            f"=IF({get_column_letter(header_map['OCC_実績'])}{row}=\"\", \"\", {get_column_letter(header_map['OCC_実績'])}{row}-{get_column_letter(header_map['OCC_FC'])}{row})"
        )
        ws.cell(row=row, column=header_map["差_ADR_実績-FC"]).value = (
            f"=IF({get_column_letter(header_map['ADR_実績'])}{row}=\"\", \"\", {get_column_letter(header_map['ADR_実績'])}{row}-{get_column_letter(header_map['ADR_FC'])}{row})"
        )
        ws.cell(row=row, column=header_map["差_売上_実績-FC"]).value = (
            f"=IF({get_column_letter(header_map['宿泊売上_実績'])}{row}=\"\", \"\", {get_column_letter(header_map['宿泊売上_実績'])}{row}-{get_column_letter(header_map['宿泊売上_FC'])}{row})"
        )

        for diff_col in ["差_OCC_FC-予算", "差_OCC_OH-FC", "差_OCC_実績-FC"]:
            ws.cell(row=row, column=header_map[diff_col]).number_format = "0.0%"
        for diff_col in ["差_ADR_FC-予算", "差_ADR_OH-FC", "差_ADR_実績-FC"]:
            ws.cell(row=row, column=header_map[diff_col]).number_format = "#,##0"
        for diff_col in ["差_売上_FC-予算", "差_売上_OH-FC", "差_売上_実績-FC"]:
            ws.cell(row=row, column=header_map[diff_col]).number_format = "#,##0"
        for diff_col in [
            "差_OCC_FC-予算",
            "差_OCC_OH-FC",
            "差_ADR_FC-予算",
            "差_ADR_OH-FC",
            "差_売上_FC-予算",
            "差_売上_OH-FC",
            "差_OCC_実績-FC",
            "差_ADR_実績-FC",
            "差_売上_実績-FC",
        ]:
            ws.cell(row=row, column=header_map[diff_col]).fill = PatternFill(
                start_color="FFFAD0",
                end_color="FFFAD0",
                fill_type="solid",
            )

        # 曜日装飾
        w_cell = ws.cell(row=row, column=header_map["曜日"])
        date_cell = ws.cell(row=row, column=header_map["日付"])
        try:
            d_obj = datetime.datetime.strptime(str(date_cell.value), "%Y/%m/%d").date()
        except Exception:
            continue
        if jpholiday.is_holiday(d_obj) or w_cell.value in ["日", "祝"]:
            w_cell.fill = sun_fill
            w_cell.font = sun_font
        elif w_cell.value == "土":
            w_cell.fill = sat_fill
            w_cell.font = sat_font

    # 実績入力時にFC列とOH列をグレー化
    actual_col = get_column_letter(header_map["室数_実績"])
    formula = f"LEN(${actual_col}2)>0"
    for offset in range(0, 7):
        col_letter = get_column_letter(header_map["室数_FC"] + offset)
        rule = FormulaRule(formula=[formula], fill=gray_fill)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{data_end_row}", rule
        )
    for offset in range(0, 7):
        col_letter = get_column_letter(header_map["室数_OH"] + offset)
        rule = FormulaRule(formula=[formula], fill=gray_fill)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{data_end_row}", rule
        )

    # 合計行
    total_row = data_end_row + 1
    days_count = data_end_row - 1
    ws.cell(row=total_row, column=1, value="合計")
    for kind in ["予算", "FC", "OH", "実績"]:
        r_col = header_map[f"室数_{kind}"]
        p_col = header_map[f"人数_{kind}"]
        s_col = header_map[f"宿泊売上_{kind}"]
        occ_col = header_map[f"OCC_{kind}"]
        adr_col = header_map[f"ADR_{kind}"]
        dor_col = header_map[f"DOR_{kind}"]
        rev_col = header_map[f"RevPAR_{kind}"]
        rl = get_column_letter(r_col)
        pl = get_column_letter(p_col)
        sl = get_column_letter(s_col)
        if kind == "OH":
            act_r = get_column_letter(header_map["室数_実績"])
            act_p = get_column_letter(header_map["人数_実績"])
            act_s = get_column_letter(header_map["宿泊売上_実績"])
            ws.cell(row=total_row, column=r_col).value = (
                f"=SUM({act_r}2:{act_r}{data_end_row})+SUMIFS({rl}2:{rl}{data_end_row},{act_r}2:{act_r}{data_end_row},\"\")"
            )
            ws.cell(row=total_row, column=p_col).value = (
                f"=SUM({act_p}2:{act_p}{data_end_row})+SUMIFS({pl}2:{pl}{data_end_row},{act_p}2:{act_p}{data_end_row},\"\")"
            )
            ws.cell(row=total_row, column=s_col).value = (
                f"=SUM({act_s}2:{act_s}{data_end_row})+SUMIFS({sl}2:{sl}{data_end_row},{act_s}2:{act_s}{data_end_row},\"\")"
            )
        else:
            ws.cell(row=total_row, column=r_col).value = f"=SUM({rl}2:{rl}{data_end_row})"
            ws.cell(row=total_row, column=p_col).value = f"=SUM({pl}2:{pl}{data_end_row})"
            ws.cell(row=total_row, column=s_col).value = f"=SUM({sl}2:{sl}{data_end_row})"
        ws.cell(row=total_row, column=occ_col).value = (
            f"=IF(COUNT({rl}2:{rl}{data_end_row})=0, \"\", SUM({rl}2:{rl}{data_end_row})/{capacity}/COUNT({rl}2:{rl}{data_end_row}))"
        )
        ws.cell(row=total_row, column=adr_col).value = (
            f"=IF(COUNT({sl}2:{sl}{data_end_row})=0, \"\", {sl}{total_row}/{rl}{total_row})"
        )
        ws.cell(row=total_row, column=dor_col).value = (
            f"=IF(COUNT({pl}2:{pl}{data_end_row})=0, \"\", {pl}{total_row}/{rl}{total_row})"
        )
        ws.cell(row=total_row, column=rev_col).value = (
            f"=IF(COUNT({sl}2:{sl}{data_end_row})=0, \"\", {sl}{total_row}/{capacity}/{days_count})"
        )
        for col in [r_col, p_col, s_col, occ_col, adr_col, dor_col, rev_col]:
            ws.cell(row=total_row, column=col).number_format = ws.cell(row=2, column=col).number_format

    for diff_col, l, r in [
        ("差_OCC_FC-予算", "OCC_FC", "OCC_予算"),
        ("差_ADR_FC-予算", "ADR_FC", "ADR_予算"),
        ("差_売上_FC-予算", "宿泊売上_FC", "宿泊売上_予算"),
        ("差_OCC_OH-FC", "OCC_OH", "OCC_FC"),
        ("差_ADR_OH-FC", "ADR_OH", "ADR_FC"),
        ("差_売上_OH-FC", "宿泊売上_OH", "宿泊売上_FC"),
        ("差_OCC_実績-FC", "OCC_実績", "OCC_FC"),
        ("差_ADR_実績-FC", "ADR_実績", "ADR_FC"),
        ("差_売上_実績-FC", "宿泊売上_実績", "宿泊売上_FC"),
    ]:
        left_col = header_map[l]
        right_col = header_map[r]
        ltr = get_column_letter(left_col)
        rtr = get_column_letter(right_col)
        if diff_col in ["差_売上_FC-予算", "差_売上_実績-FC"]:
            range_left = f"{ltr}2:{ltr}{data_end_row}"
            range_right = f"{rtr}2:{rtr}{data_end_row}"
            ws.cell(row=total_row, column=header_map[diff_col]).value = (
                f"=IF(COUNT({range_left})=0, \"\", SUMIF({range_left}, \"<>\", {range_left})-SUMIF({range_left}, \"<>\", {range_right}))"
            )
        else:
            ws.cell(row=total_row, column=header_map[diff_col]).value = (
                f"=IF({ltr}{total_row}=\"\", \"\", {ltr}{total_row}-{rtr}{total_row})"
            )
        ws.cell(row=total_row, column=header_map[diff_col]).number_format = ws.cell(row=2, column=header_map[diff_col]).number_format
        ws.cell(row=total_row, column=header_map[diff_col]).fill = PatternFill(
            start_color="FFFAD0",
            end_color="FFFAD0",
            fill_type="solid",
        )
    # 修正月次フォーキャスト
    forecast_row = total_row + 1
    ws.cell(row=forecast_row, column=1, value="修正月次フォーキャスト")
    metrics = ["室数", "人数", "宿泊売上"]
    for m in metrics:
        fc_col = header_map[f"{m}_FC"]
        act_col = header_map[f"{m}_実績"]
        fc_letter = get_column_letter(fc_col)
        act_letter = get_column_letter(act_col)
        cell = ws.cell(row=forecast_row, column=fc_col)
        cell.value = (
            f"=SUM({act_letter}2:{act_letter}{data_end_row})+SUMIFS({fc_letter}2:{fc_letter}{data_end_row},{act_letter}2:{act_letter}{data_end_row},\"\")"
        )
        cell.number_format = ws.cell(row=2, column=fc_col).number_format

    days_count = data_end_row - 1
    ws.cell(row=forecast_row, column=header_map["OCC_FC"]).value = (
        f"=IF({get_column_letter(header_map['室数_FC'])}{forecast_row}=\"\", \"\", {get_column_letter(header_map['室数_FC'])}{forecast_row}/({capacity}*{days_count}))"
    )
    ws.cell(row=forecast_row, column=header_map["ADR_FC"]).value = (
        f"=IF(OR({get_column_letter(header_map['宿泊売上_FC'])}{forecast_row}=\"\", {get_column_letter(header_map['室数_FC'])}{forecast_row}=\"\"), \"\", {get_column_letter(header_map['宿泊売上_FC'])}{forecast_row}/{get_column_letter(header_map['室数_FC'])}{forecast_row})"
    )
    ws.cell(row=forecast_row, column=header_map["DOR_FC"]).value = (
        f"=IF(OR({get_column_letter(header_map['人数_FC'])}{forecast_row}=\"\", {get_column_letter(header_map['室数_FC'])}{forecast_row}=\"\"), \"\", {get_column_letter(header_map['人数_FC'])}{forecast_row}/{get_column_letter(header_map['室数_FC'])}{forecast_row})"
    )
    ws.cell(row=forecast_row, column=header_map["RevPAR_FC"]).value = (
        f"=IF({get_column_letter(header_map['宿泊売上_FC'])}{forecast_row}=\"\", \"\", {get_column_letter(header_map['宿泊売上_FC'])}{forecast_row}/({capacity}*{days_count}))"
    )
    for m in ["OCC", "ADR", "DOR", "RevPAR"]:
        ws.cell(row=forecast_row, column=header_map[f"{m}_FC"]).number_format = ws.cell(row=2, column=header_map[f"{m}_FC"]).number_format

    # 背景色設定
    all_metrics = ["室数", "人数", "宿泊売上", "OCC", "ADR", "DOR", "RevPAR"]
    budget_cols = [header_map[f"{m}_予算"] for m in all_metrics if f"{m}_予算" in header_map]
    fc_cols = [header_map[f"{m}_FC"] for m in all_metrics if f"{m}_FC" in header_map]
    for col in budget_cols:
        for r in range(2, forecast_row + 1):
            ws.cell(row=r, column=col).fill = budget_fill
    for col in fc_cols:
        for r in range(2, forecast_row + 1):
            ws.cell(row=r, column=col).fill = fc_fill

    diff_cols = [
        header_map["差_OCC_FC-予算"],
        header_map["差_ADR_FC-予算"],
        header_map["差_売上_FC-予算"],
        header_map["差_OCC_OH-FC"],
        header_map["差_ADR_OH-FC"],
        header_map["差_売上_OH-FC"],
        header_map["差_OCC_実績-FC"],
        header_map["差_ADR_実績-FC"],
        header_map["差_売上_実績-FC"],
    ]
    for col in diff_cols:
        for r in range(2, forecast_row + 1):
            ws.cell(row=r, column=col).fill = PatternFill(
                start_color="FFFAD0",
                end_color="FFFAD0",
                fill_type="solid",
            )
        col_letter = get_column_letter(col)
        neg_rule = FormulaRule(
            formula=[f"AND(ISNUMBER({col_letter}2),{col_letter}2<0)"],
            font=Font(color="FF0000"),
        )
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{forecast_row}",
            neg_rule,
        )

    summary_dict[(year, month)] = {
        "sheet": ws.title,
        "total_row": total_row,
        "header_map": header_map,
        "data_end_row": data_end_row,
    }

    # 罫線設定
    block_ends = [
        header_map["RevPAR_予算"],
        header_map["RevPAR_FC"],
        header_map["RevPAR_OH"],
        header_map["RevPAR_実績"],
    ]
    for r in ws.iter_rows(min_row=1, max_row=forecast_row, max_col=ws.max_column):
        for c in r:
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for end_col in block_ends:
        for row in range(1, forecast_row + 1):
            cell = ws.cell(row=row, column=end_col)
            cell.border = Border(
                top=cell.border.top,
                bottom=cell.border.bottom,
                left=cell.border.left,
                right=medium,
            )
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.border = Border(top=medium, bottom=medium, left=cell.border.left, right=cell.border.right)
    for row_idx in [total_row, forecast_row]:
        for cell in ws[row_idx]:
            cell.border = Border(top=medium, bottom=medium, left=cell.border.left, right=cell.border.right)



# === 年間集計シート ===
summary = wb.create_sheet(title="年間集計")
metrics = ["室数", "人数", "宿泊売上", "OCC", "ADR", "DOR", "RevPAR"]
kinds = ["予算", "FC", "OH", "実績"]
month_labels: list[str] = []
month_keys: list[tuple[int, int]] = []
days: list[int] = []
match = re.search(r"(20\d{2})", file_path)
start_year = int(match.group(1)) if match else datetime.date.today().year
year = start_year
month = start_month
for _ in range(12):
    month_labels.append(f"{year}年{month}月")
    month_keys.append((year, month))
    info = summary_dict.get((year, month))
    days.append(info["data_end_row"] - 1 if info else 0)
    if month == 12:
        month = 1
        year += 1
    else:
        month += 1

summary.freeze_panes = "B3"
summary_totals: dict[str, dict[str, str]] = {}
current_row = 1
for kind in kinds:
    summary.cell(row=current_row, column=1, value=kind).font = Font(bold=True)
    header_row = current_row + 1
    summary.cell(row=header_row, column=1, value="指標")
    for idx, label in enumerate(month_labels, start=2):
        summary.cell(row=header_row, column=idx, value=label)
    total_col = len(month_labels) + 2
    summary.cell(row=header_row, column=total_col, value="年間合計")
    metric_rows: dict[str, int] = {}
    for metric_idx, metric in enumerate(metrics, start=header_row + 1):
        metric_rows[metric] = metric_idx
        summary.cell(row=metric_idx, column=1, value=metric)
        for m_idx, (y, m) in enumerate(month_keys, start=2):
            cell = summary.cell(row=metric_idx, column=m_idx)
            info = summary_dict.get((y, m))
            if info:
                sheet = wb[info["sheet"]]
                hmap = info["header_map"]
                tr = info["total_row"]
                col = hmap.get(f"{metric}_{kind}")
                if col:
                    cell.value = f"='{sheet.title}'!{get_column_letter(col)}{tr}"
                else:
                    cell.value = 0
            else:
                cell.value = 0
            cell.number_format = {
                "室数": "#,##0",
                "人数": "#,##0",
                "宿泊売上": "#,##0",
                "OCC": "0.0%",
                "ADR": "#,##0",
                "DOR": "0.00",
                "RevPAR": "#,##0",
            }[metric]
        col_start = get_column_letter(2)
        col_end = get_column_letter(total_col - 1)
        total_cell = summary.cell(row=metric_idx, column=total_col)
        if metric in ["室数", "人数", "宿泊売上", "DOR"]:
            total_cell.value = f"=SUM({col_start}{metric_idx}:{col_end}{metric_idx})"
        if metric in ["室数", "人数", "宿泊売上"]:
            total_cell.number_format = "#,##0"
        elif metric == "DOR":
            total_cell.number_format = "0.00"
        summary_totals.setdefault(kind, {})[metric] = total_cell.coordinate
    days_sum = sum(days) if days else 0
    room_total = summary_totals[kind]["室数"]
    pax_total = summary_totals[kind]["人数"]
    sales_total = summary_totals[kind]["宿泊売上"]
    summary.cell(row=metric_rows["OCC"], column=total_col).value = (
        f"=IFERROR({room_total}/{capacity}/{days_sum}, \"\")"
    )
    summary.cell(row=metric_rows["ADR"], column=total_col).value = (
        f"=IFERROR({sales_total}/{room_total}, \"\")"
    )
    summary.cell(row=metric_rows["DOR"], column=total_col).value = (
        f"=IFERROR({pax_total}/{room_total}, \"\")"
    )
    summary.cell(row=metric_rows["RevPAR"], column=total_col).value = (
        f"=IFERROR({sales_total}/{capacity}/{days_sum}, \"\")"
    )
    for metric in ["OCC", "ADR", "DOR", "RevPAR"]:
        summary.cell(row=metric_rows[metric], column=total_col).number_format = {
            "OCC": "0.0%",
            "ADR": "#,##0",
            "DOR": "0.00",
            "RevPAR": "#,##0",
        }[metric]
    block_end = header_row + len(metrics)
    for r in range(header_row, block_end + 1):
        for c in range(1, total_col + 1):
            summary.cell(row=r, column=c).fill = (
                budget_fill if kind == "予算" else fc_fill if kind == "FC" else PatternFill()
            )
    for cell in summary[header_row]:
        cell.font = Font(bold=True)
    for r in range(header_row - 1, block_end + 1):
        for c in range(1, total_col + 1):
            cell = summary.cell(row=r, column=c)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for c in range(1, total_col + 1):
        cell = summary.cell(row=block_end, column=c)
        cell.border = Border(top=cell.border.top, bottom=medium, left=cell.border.left, right=cell.border.right)
    current_row = block_end + 2
# === 年間差異シート ===
variance = wb.create_sheet(title="年間差異")
blocks = [
    ("FC − 予算", "FC", "予算", PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")),
    ("実績 − 予算", "実績", "予算", PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")),
    ("OH − FC", "OH", "FC", PatternFill(start_color="DDFFDD", end_color="DDFFDD", fill_type="solid")),
    ("実績 − FC", "実績", "FC", PatternFill(start_color="DDFFDD", end_color="DDFFDD", fill_type="solid")),
]
for title, left, right, fill in blocks:
    start_row = variance.max_row + 1
    variance.cell(row=start_row, column=1, value=title).font = Font(bold=True)
    header_row = start_row + 1
    variance.cell(row=header_row, column=1, value="指標")
    for idx, label in enumerate(month_labels, start=2):
        variance.cell(row=header_row, column=idx, value=label)
    total_col = len(month_labels) + 2
    variance.cell(row=header_row, column=total_col, value="年間合計")
    metric_rows: dict[str, int] = {}
    for metric_idx, metric in enumerate(metrics, start=header_row + 1):
        metric_rows[metric] = metric_idx
        variance.cell(row=metric_idx, column=1, value=metric)
        for m_idx, (y, m) in enumerate(month_keys, start=2):
            cell = variance.cell(row=metric_idx, column=m_idx)
            info = summary_dict.get((y, m))
            if info:
                sheet = wb[info["sheet"]]
                tr = info["total_row"]
                hmap = info["header_map"]
                l_col = hmap.get(f"{metric}_{left}")
                r_col = hmap.get(f"{metric}_{right}")
                if l_col and r_col:
                    l_addr = f"'{sheet.title}'!{get_column_letter(l_col)}{tr}"
                    r_addr = f"'{sheet.title}'!{get_column_letter(r_col)}{tr}"
                    if metric in ["室数", "人数", "宿泊売上"]:
                        base_formula = f"IF(OR({l_addr}=\"\", {l_addr}=0), \"\", {l_addr}-{r_addr})"
                    else:
                        base_formula = f"IF({l_addr}=\"\", \"\", {l_addr}-{r_addr})"
                    cell.value = f"=IFERROR({base_formula}, \"\")"
            cell.number_format = {
                "室数": "#,##0",
                "人数": "#,##0",
                "宿泊売上": "#,##0",
                "OCC": "0.0%",
                "ADR": "#,##0",
                "DOR": "0.00",
                "RevPAR": "#,##0",
            }[metric]
    col_start = get_column_letter(2)
    col_end = get_column_letter(total_col - 1)
    for metric in metrics:
        row = metric_rows[metric]
        cell = variance.cell(row=row, column=total_col)
        if metric in ["室数", "人数", "宿泊売上", "DOR"]:
            cell.value = f"=SUM({col_start}{row}:{col_end}{row})"
        elif metric == "OCC":
            l_room = summary_totals[left]["室数"]
            r_room = summary_totals[right]["室数"]
            cell.value = f"=IFERROR('年間集計'!{l_room}/{capacity}/{sum(days)}-'年間集計'!{r_room}/{capacity}/{sum(days)}, \"\")"
        elif metric == "ADR":
            l_room = summary_totals[left]["室数"]
            r_room = summary_totals[right]["室数"]
            l_sales = summary_totals[left]["宿泊売上"]
            r_sales = summary_totals[right]["宿泊売上"]
            cell.value = f"=IFERROR('年間集計'!{l_sales}/'年間集計'!{l_room}-'年間集計'!{r_sales}/'年間集計'!{r_room}, \"\")"
        elif metric == "RevPAR":
            l_sales = summary_totals[left]["宿泊売上"]
            r_sales = summary_totals[right]["宿泊売上"]
            cell.value = f"=IFERROR('年間集計'!{l_sales}/{capacity}/{sum(days)}-'年間集計'!{r_sales}/{capacity}/{sum(days)}, \"\")"
        cell.number_format = {
            "室数": "#,##0",
            "人数": "#,##0",
            "宿泊売上": "#,##0",
            "OCC": "0.0%",
            "ADR": "#,##0",
            "DOR": "0.00",
            "RevPAR": "#,##0",
        }[metric]
    block_end = header_row + len(metrics)
    for r in range(header_row, block_end + 1):
        for c in range(1, total_col + 1):
            variance.cell(row=r, column=c).fill = fill
            variance.cell(row=r, column=c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for c in range(1, total_col + 1):
        cell = variance.cell(row=header_row, column=c)
        cell.border = Border(top=medium, bottom=cell.border.bottom, left=cell.border.left, right=cell.border.right)
        cell = variance.cell(row=block_end, column=c)
        cell.border = Border(top=cell.border.top, bottom=medium, left=cell.border.left, right=cell.border.right)
    for idx in range(2, total_col + 1):
        col_letter = get_column_letter(idx)
        neg_rule = FormulaRule(
            formula=[f"AND(ISNUMBER({col_letter}{header_row + 1}),{col_letter}{header_row + 1}<0)"],
            font=Font(color="FF0000"),
        )
        variance.conditional_formatting.add(
            f"{col_letter}{header_row + 1}:{col_letter}{block_end}", neg_rule
        )

# === 保存 ===
match = re.search(r"(20\d{2})", file_path)
year_str = match.group(1) if match else str(datetime.date.today().year)
out_path = f"予実管理表_{year_str}年度.xlsx"
wb.save(out_path)
print(f"✅ 出力完了: {out_path}")

